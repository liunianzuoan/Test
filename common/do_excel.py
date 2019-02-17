from openpyxl import load_workbook
from common import contants


class Cases:
    def __init__(self):   # 利用对象来访问它的属性
        self.case_id = None
        self.title = None
        self.method = None
        self.url = None
        self.data = None
        self.expected = None


class DoExcel:
    def __init__(self, file_name):
        # 操作的文件
        self.file_name = file_name
        try:
            self.wb = load_workbook(self.file_name)   # 实例化一个workbook的对象
        except Exception as e:    # 异常处理
            print('错误是：{},找不到该文件'.format(e))

    def get_data(self, sheet_name):
        self.sheet_name = sheet_name
        sheet = self.wb[sheet_name]    # 获取sheet表单
        case =[]     # 列表来存在所有的数据
        for i in range(2, sheet.max_row+1):
            row_case = Cases()   # 实例化一个对象，来访问它的属性
            row_case.case_id = sheet.cell(i, 1).value    # 存在每一行的case_id
            row_case.title = sheet.cell(i, 2).value    # 存放每一行的title
            row_case.method = sheet.cell(i, 3).value    # 存放每一行的method
            row_case.url = sheet.cell(i, 4).value    # 存放每一行的url
            row_case.data = sheet.cell(i, 5).value    # 存放每一行的测试数据data
            row_case.expected = sheet.cell(i, 6).value    # 存放每一行的expected
            case.append(row_case)
        return case       # for循环结束后返回这个列表

    def write_back(self, row, actual, result):
        sheet = self.wb[self.sheet_name]    # 获取sheet表单
        sheet.cell(row, 7).value = actual    # 写入实际结果
        sheet.cell(row, 8).value = result     # 写入执行结果
        self.wb.save(self.file_name)     # 保存，记得关闭


if __name__ == '__main__':
    do_excel = DoExcel(contants.data_file)
    cases = do_excel.get_data('register')
    for case in cases:
        print(case.data)
