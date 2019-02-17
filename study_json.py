from openpyxl import load_workbook
from common import contants



class Cases:
    def __init__(self):  #利用对象来访问它的属性
        self.case_id=None
        self.title=None
        self.method = None
        self.url=None
        self.data=None
        self.expected=None

class DoExcel:
    def __init__(self,file_name):
        #操作的文件
        self.file_name=file_name
        try:
            self.wb=load_workbook(self.file_name) #实例化一个workbook的对象
        except Exception as e:  #异常处理
            print('错误是：{},找不到该文件'.format(e))

    def get_data(self,sheet_name):
        self.sheet_name=sheet_name
        sheet=self.wb[sheet_name]  #获取sheet表单
        case=[]  #列表来存在所有的数据
        for i in range(2,sheet.max_row+1):
            row_case=Cases()  #实例化一个对象，来访问它的属性
            row_case.case_id=sheet.cell(i,1).value  #存在每一行的case_id
            row_case.title=sheet.cell(i,2).value  #存放每一行的title
            row_case.method=sheet.cell(i,3).value #存放每一行的method
            row_case.url=sheet.cell(i,4).value  #存放每一行的url
            row_case.data=sheet.cell(i,5).value  #存放每一行的测试数据data
            row_case.expected=sheet.cell(i,6).value #存放每一行的expected
            case.append(row_case)  #每一行的数据存放在一个对象里面
        return case  #for循环结束后返回这个列表

    def write_back(self,row,actual,result):
        sheet=self.wb[self.sheet_name] #获取sheet表单
        sheet.cell(row,7).value=actual  #写入实际结果
        sheet.cell(row,8).value=result  #写入执行结果
        self.wb.save(self.file_name)  #保存，记得关闭



if __name__ == '__main__':
    # data=DoExcel(ProjectPath().data()).get_data('register')
    # print(data)
    # print()
    #do_excel=DoExcel(r'../datas/data.xlsx')
    do_excel=DoExcel(contants.data_file)
    cases=do_excel.get_data('register')
    from common.request import Request
    request=Request()#实例化一个对象
    for case in cases:
        print(type(case.data))
        res=request.request(case.method,case.url,case.data)
        print(type(res.text),res.text)
        print(type(res.json()),res.json())
        if res.text==case.expected:
            do_excel.write_back(case.case_id+1,res.text,'PASS')
        else:
            do_excel.write_back(case.case_id + 1, res.text, 'Failed')



'''
json  json库  
json数据格式转换成相应的语言里面的数据类型  
json转成python里面的dict


'''

import json
people='{"name":"lily","age":18,"married":false}'
p_dict=json.loads(people) #利用json.loads()转换成字典
#json序列化 将字符串转成字典
print(type(p_dict),p_dict)

#json序列化，将file转成字典
fp=open(contants.json_file)
f_dict=json.load(fp=fp)
print(f_dict['name'])



